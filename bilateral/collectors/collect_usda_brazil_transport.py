#!/usr/bin/env python3
"""
collect_usda_brazil_transport.py — AgriMacro Intelligence
Coleta dados do USDA AMS Brazil Soybean Transportation Guide
Fonte: https://www.ams.usda.gov/services/transportation-analysis/soybean-datasets

Datasets coletados:
  1. Cost of transporting BR soybeans from southern ports → Shanghai (2006-24)
  2. Cost of transporting BR soybeans from northern ports → Shanghai (2015-24)
  3. Quarterly ocean freight rates BR ports → Germany/China (2005-24)
  4. Monthly truck transportation cost index (2003-24)
  5. Quarterly truck rates by route (2024)

Output: JSON padronizado AgriMacro em data/usda_brazil_transport/
Uso bilateral: Lado Brasil do "Landed Cost Spread Shanghai"
Principio: ZERO MOCK — apenas dados reais do USDA AMS / ESALQ
"""

import json
import logging
import os
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path

import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

BASE_URL = "https://www.ams.usda.gov/sites/default/files/media"

DATASETS = {
    "southern_shanghai": {
        "url": f"{BASE_URL}/CostoftransportingBraziliansoybeansfromthesouthernportstoShanghaiChina_2006_24.xlsx",
        "sheet": "Table 1",
        "label": "Southern Ports → Shanghai (landed cost)",
    },
    "northern_shanghai": {
        "url": f"{BASE_URL}/CostoftransportingBraziliansoybeansfromthenorthernandnortheasternportstoShanghaiChina_2015_24.xlsx",
        "sheet": 0,  # first sheet (name varies)
        "label": "Northern Ports → Shanghai (landed cost)",
    },
    "ocean_freight": {
        "url": f"{BASE_URL}/QuarterlyoceanfreightratesforshippingsoybeansfromselectedBrazilianportstoGermanyandChina_2005_24.xlsx",
        "sheet": "Table 9",
        "label": "Quarterly Ocean Freight Rates (BR → Germany/China)",
    },
    "truck_index": {
        "url": f"{BASE_URL}/MonthlyBraziliansoybeanexporttrucktransportationcostindex_2005_24.xlsx",
        "sheet": "Table 8",
        "label": "Monthly Truck Cost Index",
    },
    "truck_routes": {
        "url": f"{BASE_URL}/QuarterlyTruckRates_2024.xlsx",
        "sheet": "Table 7",
        "label": "Quarterly Truck Rates by Route (2024)",
    },
}

BASE_DIR = Path(os.environ.get("AGRIMACRO_DATA_DIR", "data"))
OUTPUT_DIR = BASE_DIR / "usda_brazil_transport"
CACHE_DIR = OUTPUT_DIR / "cache"
RAW_DIR = OUTPUT_DIR / "raw_xlsx"

LOG_FORMAT = "%(asctime)s [%(levelname)s] %(message)s"
logging.basicConfig(level=logging.INFO, format=LOG_FORMAT)
logger = logging.getLogger("collect_usda_brazil")


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def download_xlsx(url, save_path=None, timeout=60):
    logger.info(f"Downloading: {url}")
    resp = requests.get(url, timeout=timeout, verify=False)
    resp.raise_for_status()
    if save_path:
        Path(save_path).parent.mkdir(parents=True, exist_ok=True)
        with open(save_path, "wb") as f:
            f.write(resp.content)
    logger.info(f"  {len(resp.content):,} bytes")
    return resp.content


def open_sheet(content, sheet_ref):
    """Open xlsx sheet. sheet_ref can be name (str) or index (int)."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    if isinstance(sheet_ref, int):
        ws = wb[wb.sheetnames[sheet_ref]]
    else:
        ws = wb[sheet_ref]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _j(val):
    """Convert to JSON-safe type."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        return round(val, 4) if isinstance(val, float) else val
    s = str(val).strip()
    if s in ("", "-", "  ", "n/a", "N/A", "--"):
        return None
    try:
        return round(float(s.replace(",", "")), 4)
    except ValueError:
        return s


# ─────────────────────────────────────────────
# PARSER: Landed Cost tables (Southern & Northern)
# ─────────────────────────────────────────────

def parse_landed_cost(content, sheet_ref):
    """
    Parse landed cost tables (Southern→Shanghai, Northern→Shanghai).
    
    Structure: Multiple route blocks within one sheet.
    Each block:
      Row: route title (e.g., "North MT - Santos by truck")
      Row: "US$/mt--"
      Row: years (2006, 2007, ..., 2024)
      Row: blank
      Row: Truck | values...
      Row: Ocean | values...
      Row: Total transportation | values...
      Row: Farm gate price | values...
      Row: Landed cost | values...
      Row: Transport % of landed cost | values...
    """
    rows = open_sheet(content, sheet_ref)
    
    routes = []
    i = 0
    
    while i < len(rows):
        row = rows[i]
        
        # Detect route title: typically row[1] contains long string with route name
        # followed by years row and then data rows
        if (row and len(row) > 1 and row[1] and isinstance(row[1], str) 
            and any(kw in str(row[1]).lower() for kw in ["mt", "go", "pr", "rs", "pi", "ma", "ba", "to", "santos", "paranaguá", "rio grande", "santarém", "são luís", "barcarena", "salvador"])):
            
            route_name = str(row[1]).strip()
            
            # Find years row (next row with numeric values)
            years = []
            j = i + 1
            while j < len(rows) and not years:
                candidate = rows[j]
                if candidate and len(candidate) > 2:
                    nums = [v for v in candidate[1:] if isinstance(v, (int, float)) and 2000 <= v <= 2030]
                    if len(nums) >= 3:
                        years = [int(v) for v in candidate[1:] if isinstance(v, (int, float)) and 2000 <= v <= 2030]
                j += 1
            
            if not years:
                i += 1
                continue
            
            # Parse data rows after years (skip blank row)
            data_start = j  # j is already past years row
            # Skip blank row if present
            if data_start < len(rows) and all(v is None for v in (rows[data_start] or [])):
                data_start += 1
            
            components = {}
            COMPONENT_KEYS = {
                "truck": "truck_usd_mt",
                "rail": "rail_usd_mt",
                "barge": "barge_usd_mt",
                "ocean": "ocean_usd_mt",
                "total transportation": "total_transport_usd_mt",
                "farm gate": "farm_gate_usd_mt",
                "landed cost": "landed_cost_usd_mt",
                "transport %": "transport_pct_landed",
            }
            
            for k in range(data_start, min(data_start + 8, len(rows))):
                if k >= len(rows):
                    break
                drow = rows[k]
                if not drow or not drow[0]:
                    continue
                label = str(drow[0]).strip().lower()
                
                matched_key = None
                for pattern, key in COMPONENT_KEYS.items():
                    if pattern in label:
                        matched_key = key
                        break
                
                if matched_key:
                    values = {}
                    for yi, year in enumerate(years):
                        col_idx = yi + 1
                        if col_idx < len(drow):
                            values[str(year)] = _j(drow[col_idx])
                    components[matched_key] = values
            
            if components:
                routes.append({
                    "route": route_name,
                    "years": years,
                    "components": components,
                })
        
        i += 1
    
    logger.info(f"  Landed cost: {len(routes)} routes parsed")
    return routes


# ─────────────────────────────────────────────
# PARSER: Ocean Freight
# ─────────────────────────────────────────────

def parse_ocean_freight(content):
    """
    Sheet "Table 9": Repeating blocks per year.
    Each block:
      Row: Port | Destination | 1st qtr YYYY | 2nd qtr YYYY | 3rd qtr YYYY | 4th qtr YYYY | Average YYYY
      Row: blank
      Row: Santos | Germany (Hamburg) | values...
      Row: Paranaguá | Germany (Hamburg) | values...
      Row: Rio Grande | Germany (Hamburg) | values...
      Row: Santos | China (Shanghai) | values...
      Row: Paranaguá | China (Shanghai) | values...
      Row: Rio Grande | China (Shanghai) | values...
      (+ São Luís, Barcarena, Santarém in later years)
    """
    rows = open_sheet(content, "Table 9")
    
    records = []
    current_year = None
    
    i = 0
    while i < len(rows):
        row = rows[i]
        if not row or len(row) < 8:
            i += 1
            continue
        
        # Detect header row: contains "Port" and quarter labels
        if row[1] and str(row[1]).strip() == "Port":
            # Extract year from quarter headers
            for col in row[3:7]:
                if col and isinstance(col, str):
                    for word in str(col).split():
                        try:
                            yr = int(word)
                            if 2000 <= yr <= 2030:
                                current_year = yr
                                break
                        except ValueError:
                            pass
                if current_year:
                    break
            i += 1
            continue
        
        # Data row: port name in col 1, destination in col 2
        port = row[1]
        dest = row[2]
        
        if (port and dest and current_year 
            and isinstance(port, str) and isinstance(dest, str)
            and any(p in port for p in ["Santos", "Paranaguá", "Rio Grande", "São Luís", "Barcarena", "Santarém"])):
            
            rec = {
                "year": current_year,
                "port": port.strip(),
                "destination": dest.strip(),
                "q1_usd_mt": _j(row[3]) if len(row) > 3 else None,
                "q2_usd_mt": _j(row[4]) if len(row) > 4 else None,
                "q3_usd_mt": _j(row[5]) if len(row) > 5 else None,
                "q4_usd_mt": _j(row[6]) if len(row) > 6 else None,
                "avg_usd_mt": _j(row[7]) if len(row) > 7 else None,
            }
            records.append(rec)
        
        i += 1
    
    logger.info(f"  Ocean freight: {len(records)} port-destination-year records")
    
    # Extract latest Santos→Shanghai for summary
    latest_shanghai = None
    for rec in reversed(records):
        if "Santos" in rec["port"] and "Shanghai" in rec["destination"]:
            latest_shanghai = rec
            break
    
    return {"records": records, "latest_santos_shanghai": latest_shanghai}


# ─────────────────────────────────────────────
# PARSER: Truck Cost Index
# ─────────────────────────────────────────────

def parse_truck_index(content):
    """
    Sheet "Table 8":
      Row 5: headers (MONTH, Freight price US$/mt/100mi, Index variation %, Index value)
      Row 6+: monthly data from 2003
    """
    rows = open_sheet(content, "Table 8")
    
    records = []
    for row in rows[6:]:
        if not row or len(row) < 3:
            continue
        date_val = row[1]
        if not isinstance(date_val, datetime):
            continue
        
        rec = {
            "date": date_val.strftime("%Y-%m-%d"),
            "freight_usd_mt_100mi": _j(row[2]),
            "index_variation_pct": _j(row[3]) if len(row) > 3 else None,
            "index_value": _j(row[4]) if len(row) > 4 else None,
        }
        if rec["freight_usd_mt_100mi"] is not None:
            records.append(rec)
    
    logger.info(f"  Truck index: {len(records)} monthly records" +
                (f" ({records[0]['date']} to {records[-1]['date']})" if records else ""))
    return records


# ─────────────────────────────────────────────
# PARSER: Truck Routes
# ─────────────────────────────────────────────

def parse_truck_routes(content):
    """
    Sheet "Table 7":
      Row 0: title
      Row 1: headers (Route #, Origin, Destination, Distance, Share %, Freight price)
      Row 2: sub-headers (1st qtr, 2nd qtr, 3rd qtr, 4th qtr)
      Row 3+: route data
    """
    rows = open_sheet(content, "Table 7")
    
    records = []
    for row in rows[3:]:
        if not row or len(row) < 10:
            continue
        route_num = row[1]
        if not isinstance(route_num, (int, float)):
            continue
        
        rec = {
            "route_num": int(route_num),
            "origin": str(row[2]).strip() if row[2] else None,
            "destination": str(row[3]).strip() if row[3] else None,
            "distance_miles": _j(row[4]),
            "share_pct": _j(row[5]),
            "q1_usd_mt_100mi": _j(row[6]),
            "q2_usd_mt_100mi": _j(row[7]),
            "q3_usd_mt_100mi": _j(row[8]),
            "q4_usd_mt_100mi": _j(row[9]),
        }
        records.append(rec)
    
    logger.info(f"  Truck routes: {len(records)} routes")
    return records


# ─────────────────────────────────────────────
# SUMMARY
# ─────────────────────────────────────────────

def build_summary(parsed):
    summary = {}
    
    # Latest Santos→Shanghai ocean freight
    of = parsed.get("ocean_freight", {})
    if isinstance(of, dict) and of.get("latest_santos_shanghai"):
        lss = of["latest_santos_shanghai"]
        summary["ocean_santos_shanghai"] = lss
    
    # Latest truck cost index
    ti = parsed.get("truck_index", [])
    if ti:
        summary["truck_index_latest"] = ti[-1]
    
    # Latest landed cost from southern routes
    sr = parsed.get("southern_shanghai", [])
    if sr:
        latest_costs = {}
        for route in sr:
            comps = route.get("components", {})
            lc = comps.get("landed_cost_usd_mt", {})
            if lc:
                latest_year = max(lc.keys())
                latest_costs[route["route"][:50]] = {
                    "year": latest_year,
                    "landed_cost_usd_mt": lc[latest_year],
                }
        summary["southern_landed_costs"] = latest_costs
    
    # Truck routes summary
    tr = parsed.get("truck_routes", [])
    if tr:
        summary["truck_routes_count"] = len(tr)
        # Key route: Sorriso→Santos
        for r in tr:
            if r.get("origin") and "Sorriso" in r["origin"] and r.get("destination") and "Santos" in r["destination"]:
                summary["sorriso_santos"] = r
                break
    
    return summary


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def save_json(data, filepath):
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    logger.info(f"Saved: {filepath}")


def load_cache():
    f = CACHE_DIR / "usda_brazil_latest.json"
    if f.exists():
        try:
            return json.load(open(f, encoding="utf-8"))
        except Exception:
            pass
    return {}


def collect(datasets=None, save_raw=True):
    now = datetime.now()
    ts = now.isoformat()
    
    if datasets is None:
        datasets = list(DATASETS.keys())
    
    parsers = {
        "southern_shanghai": lambda c: parse_landed_cost(c, "Table 1"),
        "northern_shanghai": lambda c: parse_landed_cost(c, 0),
        "ocean_freight": parse_ocean_freight,
        "truck_index": parse_truck_index,
        "truck_routes": parse_truck_routes,
    }
    
    parsed = {}
    errors = {}
    
    for key in datasets:
        if key not in DATASETS:
            continue
        info = DATASETS[key]
        logger.info(f"Processing {key}: {info['label']}")
        try:
            raw_path = RAW_DIR / f"{key}.xlsx" if save_raw else None
            content = download_xlsx(info["url"], save_path=raw_path)
            parsed[key] = parsers[key](content)
        except Exception as e:
            logger.error(f"  Error on {key}: {e}")
            errors[key] = str(e)
    
    if not parsed:
        cached = load_cache()
        if cached:
            cached["status"] = "cached"
            return cached
        return {"source": "usda_brazil_transport", "status": "error", "errors": errors}
    
    summary = build_summary(parsed)
    
    serialized = {}
    for key, data in parsed.items():
        label = DATASETS[key]["label"]
        if isinstance(data, list):
            serialized[key] = {"label": label, "count": len(data), "records": data}
        else:
            serialized[key] = {"label": label, **data}
            if "records" in data:
                serialized[key]["count"] = len(data["records"])
    
    output = {
        "source": "usda_brazil_transport",
        "source_url": "https://www.ams.usda.gov/services/transportation-analysis/soybean-datasets",
        "collection_timestamp": ts,
        "status": "ok",
        "errors": errors if errors else None,
        "summary": summary,
        "data": serialized,
    }
    
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    save_json(output, OUTPUT_DIR / f"usda_brazil_{timestamp}.json")
    save_json(output, OUTPUT_DIR / "usda_brazil_latest.json")
    save_json(output, CACHE_DIR / "usda_brazil_latest.json")
    
    # Log summary
    logger.info("=" * 60)
    logger.info("USDA BRAZIL SOYBEAN TRANSPORT - SUMMARY")
    logger.info("=" * 60)
    
    if "ocean_santos_shanghai" in summary:
        oss = summary["ocean_santos_shanghai"]
        logger.info(f"  Ocean freight Santos→Shanghai ({oss['year']}):")
        logger.info(f"    Q1: ${oss.get('q1_usd_mt', '?')}/mt  Q2: ${oss.get('q2_usd_mt', '?')}/mt  Q3: ${oss.get('q3_usd_mt', '?')}/mt  Q4: ${oss.get('q4_usd_mt', '?')}/mt")
        logger.info(f"    Average: ${oss.get('avg_usd_mt', '?')}/mt")
    
    if "truck_index_latest" in summary:
        ti = summary["truck_index_latest"]
        logger.info(f"  Truck cost index ({ti['date']}): ${ti['freight_usd_mt_100mi']}/mt/100mi")
    
    if "sorriso_santos" in summary:
        ss = summary["sorriso_santos"]
        logger.info(f"  Sorriso→Santos ({ss['distance_miles']}mi): Q4=${ss.get('q4_usd_mt_100mi', '?')}/mt/100mi")
    
    if "southern_landed_costs" in summary:
        logger.info("  Southern landed costs (latest year):")
        for route, data in summary["southern_landed_costs"].items():
            logger.info(f"    {route}: ${data['landed_cost_usd_mt']}/mt ({data['year']})")
    
    for key in serialized:
        count = serialized[key].get("count", "?")
        logger.info(f"  {DATASETS[key]['label']:.<55} {count:>5} records")
    
    if errors:
        for key, err in errors.items():
            logger.warning(f"  ERROR {key}: {err}")
    
    logger.info("=" * 60)
    
    return output


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="AgriMacro - USDA Brazil Soybean Transport Collector")
    parser.add_argument("--datasets", nargs="+", default=None,
                        choices=list(DATASETS.keys()))
    parser.add_argument("--output-dir", type=str, default=None)
    args = parser.parse_args()
    
    if args.output_dir:
        OUTPUT_DIR = Path(args.output_dir)
        CACHE_DIR = OUTPUT_DIR / "cache"
        RAW_DIR = OUTPUT_DIR / "raw_xlsx"
    
    result = collect(datasets=args.datasets)
    sys.exit(0 if result["status"] in ("ok", "cached") else 1)
